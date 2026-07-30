#!/usr/bin/env python3
"""
AZ-104 question bank assembler.
Reads batch JSON files -> normalizes -> semantic dedup -> obsolescence sweep ->
assigns IDs -> emits az104_question_bank.json + az104_question_bank.xlsx

Anchor: skills measured as of April 17, 2026.
"""
import json, re, sys, glob, os
from datetime import date
from collections import Counter, defaultdict

BASE = os.path.dirname(os.path.abspath(__file__))
BATCH_DIR = os.path.join(BASE, "batches")
OUT_DIR = sys.argv[1] if len(sys.argv) > 1 else BASE

BUILD_DATE = "2026-07-16"
STUDY_GUIDE_VERSION = "Skills measured as of April 17, 2026"
STUDY_GUIDE_URL = "https://learn.microsoft.com/en-us/credentials/certifications/resources/study-guides/az-104"

DOMAINS = [
    "Manage Azure identities and governance",
    "Implement and manage storage",
    "Deploy and manage Azure compute resources",
    "Implement and manage virtual networking",
    "Monitor and maintain Azure resources",
]
WEIGHTS = {  # official min/max % from the study guide
    "Manage Azure identities and governance": (20, 25),
    "Implement and manage storage": (15, 20),
    "Deploy and manage Azure compute resources": (20, 25),
    "Implement and manage virtual networking": (15, 20),
    "Monitor and maintain Azure resources": (10, 15),
}

FIELDS = ["id", "dominio", "sotto_argomento", "tipo", "domanda",
          "opzione_a", "opzione_b", "opzione_c", "opzione_d", "opzione_e",
          "risposta_corretta", "spiegazione", "url_riferimento", "fonte", "url_fonte",
          "sintesi_commenti", "consenso_community", "stato", "difficolta", "tags",
          "is_generated", "data_verifica"]

VALID_TIPI = {"multiple_choice", "multiple_response", "yes_no_series",
              "hotspot", "drag_drop", "case_study"}
VALID_STATI = {"verificata", "contestata", "da_rivedere"}

removed = []  # {id, testo_breve, motivo}

# Questions the Fase-4 verifiers deleted before this script ever saw them: they are gone
# from the batch files, so their provenance has to be merged back in from the workflow report.
_ext = os.path.join(BASE, "removed_external.json")
if os.path.exists(_ext):
    try:
        with open(_ext, "r", encoding="utf-8") as fh:
            for r in json.load(fh):
                removed.append({"id": "",
                                "testo_breve": str(r.get("testo_breve", ""))[:120],
                                "motivo": r.get("motivo", "qualita insufficiente")})
        print(f"[load] {len(removed)} rimozioni ereditate dai verificatori di Fase 4")
    except Exception as e:
        print(f"[load] removed_external.json non leggibile: {e}")

# ---------------------------------------------------------------- terminology
TERM_FIXES = [
    (re.compile(r"\bAzure Active Directory\b", re.I), "Microsoft Entra ID"),
    (re.compile(r"\bAzure AD DS\b"), "Microsoft Entra Domain Services"),
    (re.compile(r"\bAzure AD Connect\b", re.I), "Microsoft Entra Connect"),
    (re.compile(r"\bAzure AD\b"), "Microsoft Entra ID"),
    (re.compile(r"\bAAD\b"), "Microsoft Entra ID"),
    (re.compile(r"\bMicrosoft Entra ID Connect\b"), "Microsoft Entra Connect"),
    (re.compile(r"\bMicrosoft Entra ID Domain Services\b"), "Microsoft Entra Domain Services"),
]

# fundamentally-retired tech: if the stem is ABOUT this, drop the question
OBSOLETE_PATTERNS = [
    (re.compile(r"\bclassic (deployment )?model\b|\bASM\b|\bAzure Service Manager\b", re.I), "obsoleta"),
    (re.compile(r"\bunmanaged disk", re.I), "obsoleta"),
    (re.compile(r"\bAzureRM\b|\bAdd-AzureRm|\bNew-AzureRm|\bGet-AzureRm", re.I), "obsoleta"),
    (re.compile(r"\bMicrosoft Monitoring Agent\b|\bMMA\b|\bLog Analytics agent\b", re.I), "obsoleta"),
    (re.compile(r"\bClassic (VM|storage account|virtual network)\b", re.I), "obsoleta"),
]

STOP = set("""a an the is are was were be been being of to in on for with and or not you your they
this that these those it its as at by from have has had do does did can could will would should may
which what when where who whom how if then than there their them he she his her we us our i me my""".split())


def norm_text(s):
    s = (s or "").lower()
    s = re.sub(r"https?://\S+", " ", s)
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return [w for w in s.split() if w not in STOP and len(w) > 2]


def shingles(s, k=3):
    w = norm_text(s)
    if len(w) < k:
        return set(w)
    return set(" ".join(w[i:i + k]) for i in range(len(w) - k + 1))


def jaccard(a, b):
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def apply_terms(q):
    for f in ("domanda", "spiegazione", "sotto_argomento", "tags",
              "opzione_a", "opzione_b", "opzione_c", "opzione_d", "opzione_e"):
        v = q.get(f)
        if isinstance(v, str):
            for pat, rep in TERM_FIXES:
                v = pat.sub(rep, v)
            q[f] = v
    return q


def coerce(q, batch_id):
    out = {}
    for f in FIELDS:
        out[f] = q.get(f, "")
    out["id"] = ""
    for f in ["dominio", "sotto_argomento", "tipo", "domanda", "risposta_corretta",
              "spiegazione", "url_riferimento", "fonte", "url_fonte",
              "sintesi_commenti", "consenso_community", "stato", "tags", "data_verifica",
              "opzione_a", "opzione_b", "opzione_c", "opzione_d", "opzione_e"]:
        v = out[f]
        out[f] = "" if v is None else str(v).strip()
    # difficolta
    try:
        d = int(out["difficolta"])
    except (ValueError, TypeError):
        d = 2
    out["difficolta"] = min(3, max(1, d))
    # is_generated
    ig = q.get("is_generated", True)
    if isinstance(ig, str):
        ig = ig.strip().lower() in ("true", "1", "yes", "vero")
    out["is_generated"] = bool(ig)
    # tipo
    if out["tipo"] not in VALID_TIPI:
        out["tipo"] = "multiple_choice"
    # stato
    if out["stato"] not in VALID_STATI:
        out["stato"] = "da_rivedere"
    if not out["data_verifica"]:
        out["data_verifica"] = BUILD_DATE
    if not out["url_fonte"]:
        out["url_fonte"] = out["url_riferimento"]
    if not out["fonte"]:
        out["fonte"] = "Generated (Microsoft Learn-anchored)"
    out["_batch"] = batch_id
    return out


def quality_ok(q):
    """Returns (ok, reason)."""
    if len(q["domanda"]) < 25:
        return False, "qualita insufficiente"
    if not q["risposta_corretta"]:
        return False, "qualita insufficiente"
    if len(q["spiegazione"]) < 40:
        return False, "qualita insufficiente"
    if "learn.microsoft.com" not in q["url_riferimento"]:
        return False, "qualita insufficiente"
    tipo = q["tipo"]
    opts = [q[f"opzione_{c}"] for c in "abcde"]
    filled = [o for o in opts if o]
    if tipo == "multiple_choice" or tipo == "case_study":
        if len(filled) < 3:
            return False, "qualita insufficiente"
        if not re.fullmatch(r"[a-e]", q["risposta_corretta"].strip().lower()):
            return False, "qualita insufficiente"
    elif tipo == "multiple_response":
        if len(filled) < 4:
            return False, "qualita insufficiente"
        parts = [p.strip().lower() for p in q["risposta_corretta"].split(",") if p.strip()]
        if len(parts) < 2 or not all(re.fullmatch(r"[a-e]", p) for p in parts):
            return False, "qualita insufficiente"
    elif tipo == "yes_no_series":
        parts = [p.strip().lower() for p in q["risposta_corretta"].split(",") if p.strip()]
        if len(parts) < 2 or not all(p in ("yes", "no") for p in parts):
            return False, "qualita insufficiente"
        if len([o for o in opts[:3] if o]) < 2:
            return False, "qualita insufficiente"
    elif tipo == "drag_drop":
        parts = [p.strip().lower() for p in q["risposta_corretta"].split(",") if p.strip()]
        if len(parts) < 2 or not all(re.fullmatch(r"[a-e]", p) for p in parts):
            return False, "qualita insufficiente"
        if len(set(parts)) != len(parts):
            return False, "qualita insufficiente"
    elif tipo == "hotspot":
        if len([p for p in q["risposta_corretta"].split(",") if p.strip()]) < 2:
            return False, "qualita insufficiente"
        if len(filled) < 2:
            return False, "qualita insufficiente"
    return True, ""


def obsolete_reason(q):
    hay = q["domanda"] + " " + " ".join(q[f"opzione_{c}"] for c in "abcde")
    for pat, motivo in OBSOLETE_PATTERNS:
        if pat.search(hay):
            return motivo
    return None


def brief(q, n=90):
    t = re.sub(r"\s+", " ", q["domanda"]).strip()
    return t[:n] + ("..." if len(t) > n else "")


# ---------------------------------------------------------------- load
files = sorted(glob.glob(os.path.join(BATCH_DIR, "*.json")))
raw = []
load_report = []
for path in files:
    bid = os.path.splitext(os.path.basename(path))[0]
    try:
        with open(path, "r", encoding="utf-8") as fh:
            txt = fh.read().strip()
        txt = re.sub(r"^```(?:json)?\s*|\s*```$", "", txt)
        data = json.loads(txt)
        if isinstance(data, dict):
            for k in ("questions", "domande", "items", "data"):
                if isinstance(data.get(k), list):
                    data = data[k]
                    break
        if not isinstance(data, list):
            load_report.append((bid, 0, "not a JSON array"))
            continue
        n = 0
        for q in data:
            if isinstance(q, dict) and q.get("domanda"):
                raw.append(coerce(q, bid))
                n += 1
        load_report.append((bid, n, "ok"))
    except Exception as e:
        load_report.append((bid, 0, f"ERROR {type(e).__name__}: {e}"))

print(f"[load] {len(files)} batch files, {len(raw)} raw questions")
for bid, n, st in load_report:
    if st != "ok" or n == 0:
        print(f"   !! {bid}: {n} ({st})")

# ---------------------------------------------------------------- clean
kept = []
for q in raw:
    q = apply_terms(q)
    if q["dominio"] not in DOMAINS:
        match = next((d for d in DOMAINS if d.lower()[:18] in q["dominio"].lower()), None)
        if match:
            q["dominio"] = match
        else:
            removed.append({"id": "", "testo_breve": brief(q), "motivo": "qualita insufficiente"})
            continue
    mo = obsolete_reason(q)
    if mo:
        removed.append({"id": "", "testo_breve": brief(q), "motivo": mo})
        continue
    ok, reason = quality_ok(q)
    if not ok:
        removed.append({"id": "", "testo_breve": brief(q), "motivo": reason})
        continue
    kept.append(q)

print(f"[clean] kept {len(kept)}, removed {len(removed)}")

# ---------------------------------------------------------------- dedup
STATE_RANK = {"verificata": 2, "contestata": 1, "da_rivedere": 0}


def score(q):
    return (STATE_RANK.get(q["stato"], 0), len(q["spiegazione"]),
            1 if q["sintesi_commenti"] else 0, len(q["url_riferimento"]))


for q in kept:
    q["_sh"] = shingles(q["domanda"])

# bucket by domain to keep comparisons tractable
buckets = defaultdict(list)
for q in kept:
    buckets[q["dominio"]].append(q)

survivors = []
dupes = 0
for dom, qs in buckets.items():
    qs.sort(key=score, reverse=True)  # best first -> keeps the best of each dup cluster
    acc = []
    for q in qs:
        hit = None
        for a in acc:
            if jaccard(q["_sh"], a["_sh"]) >= 0.72:
                hit = a
                break
        if hit:
            removed.append({"id": "", "testo_breve": brief(q), "motivo": "duplicato"})
            dupes += 1
        else:
            acc.append(q)
    survivors.extend(acc)

print(f"[dedup] removed {dupes} duplicates, {len(survivors)} remain")

# ---------------------------------------------------------------- rebalance to official weights
# The authoring agents overshot on some domains. Trim the weakest questions of an
# over-weight domain until it sits inside its official band. These are not bad questions,
# so they are logged with their own reason rather than mislabelled as poor quality.
def keep_score(q):
    """Higher = more worth keeping."""
    return (STATE_RANK.get(q["stato"], 0),
            1 if q["difficolta"] >= 2 else 0,      # applied/scenario beat pure recall
            0 if q["tipo"] == "multiple_choice" else 1,  # protect the rarer formats
            len(q["spiegazione"]))

trimmed = 0
guard = 0
while guard < 2000:
    guard += 1
    tot = len(survivors)
    over = []
    for d in DOMAINS:
        n = sum(1 for q in survivors if q["dominio"] == d)
        p = n / tot * 100 if tot else 0
        if p > WEIGHTS[d][1]:
            over.append((p - WEIGHTS[d][1], d))
    if not over:
        break
    over.sort(reverse=True)
    d = over[0][1]
    pool = [q for q in survivors if q["dominio"] == d]
    # drop the weakest, but never starve a sub-topic to zero
    sub_n = Counter(q["sotto_argomento"] for q in pool)
    cand = [q for q in pool if sub_n[q["sotto_argomento"]] > 1] or pool
    victim = min(cand, key=keep_score)
    survivors.remove(victim)
    removed.append({"id": "", "testo_breve": brief(victim), "motivo": "eccedenza peso dominio"})
    trimmed += 1

if trimmed:
    print(f"[rebalance] rimosse {trimmed} domande in eccedenza per rientrare nei pesi ufficiali")

# ---------------------------------------------------------------- order + ids
survivors.sort(key=lambda q: (DOMAINS.index(q["dominio"]), q["sotto_argomento"], q["_batch"]))
for i, q in enumerate(survivors, 1):
    q["id"] = f"AZ104-{i:04d}"
    q.pop("_sh", None)

final = [{f: q[f] for f in FIELDS} for q in survivors]

# ---------------------------------------------------------------- JSON
os.makedirs(OUT_DIR, exist_ok=True)
json_path = os.path.join(OUT_DIR, "az104_question_bank.json")
with open(json_path, "w", encoding="utf-8") as fh:
    json.dump(final, fh, ensure_ascii=False, indent=1)
print(f"[out] {json_path}  ({len(final)} questions)")

# ---------------------------------------------------------------- stats
total = len(final)
by_dom = Counter(q["dominio"] for q in final)
by_stato = Counter(q["stato"] for q in final)
by_tipo = Counter(q["tipo"] for q in final)
by_diff = Counter(q["difficolta"] for q in final)
by_fonte = Counter(q["fonte"] for q in final)
by_sub = Counter((q["dominio"], q["sotto_argomento"]) for q in final)
by_gen = Counter("originale generata" if q["is_generated"] else "raccolta da fonte aperta" for q in final)
by_removed = Counter(r["motivo"] for r in removed)

# ---------------------------------------------------------------- XLSX
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=12, color="1F3864")


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


# --- Domande
ws = wb.active
ws.title = "Domande"
ws.append(FIELDS)
for q in final:
    ws.append([q[f] for f in FIELDS])
style_header(ws, len(FIELDS))
ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}{ws.max_row}"
widths = {"id": 12, "dominio": 34, "sotto_argomento": 34, "tipo": 17, "domanda": 70,
          "opzione_a": 30, "opzione_b": 30, "opzione_c": 30, "opzione_d": 30, "opzione_e": 30,
          "risposta_corretta": 16, "spiegazione": 80, "url_riferimento": 45, "fonte": 30,
          "url_fonte": 40, "sintesi_commenti": 40, "consenso_community": 18, "stato": 13,
          "difficolta": 10, "tags": 28, "is_generated": 12, "data_verifica": 13}
for i, f in enumerate(FIELDS, 1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(f, 18)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# --- Statistiche
ws = wb.create_sheet("Statistiche")


def block(title, headers, rows, gap=1):
    ws.append([])
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=title).font = TITLE_FONT
    ws.append(headers)
    style_header(ws, len(headers), row=ws.max_row)
    for row in rows:
        ws.append(row)
    for _ in range(gap):
        ws.append([])


ws.append(["Banca domande AZ-104 — Statistiche"])
ws["A1"].font = Font(bold=True, size=14, color="1F3864")
ws.append([f"Riferimento: {STUDY_GUIDE_VERSION}"])
ws.append([f"Data build: {BUILD_DATE}"])
ws.append([f"Totale domande: {total}"])

rows = []
for d in DOMAINS:
    n = by_dom.get(d, 0)
    pct = (n / total * 100) if total else 0
    lo, hi = WEIGHTS[d]
    esito = "OK" if lo <= pct <= hi else ("SOTTO PESO" if pct < lo else "SOPRA PESO")
    rows.append([d, n, round(pct, 1), f"{lo}-{hi}%", esito])
block("Copertura per dominio vs pesi ufficiali",
      ["dominio", "n. domande", "% attuale", "peso ufficiale", "esito"], rows)

block("Conteggio per stato", ["stato", "n. domande", "%"],
      [[k, v, round(v / total * 100, 1) if total else 0] for k, v in by_stato.most_common()])

block("Conteggio per tipo", ["tipo", "n. domande", "%"],
      [[k, v, round(v / total * 100, 1) if total else 0] for k, v in by_tipo.most_common()])

block("Conteggio per difficolta", ["difficolta", "n. domande", "%"],
      [[f"{k} ({['','base','applicativa','scenario complesso'][k]})", v,
        round(v / total * 100, 1) if total else 0] for k, v in sorted(by_diff.items())])

block("Conteggio per origine", ["origine", "n. domande", "%"],
      [[k, v, round(v / total * 100, 1) if total else 0] for k, v in by_gen.most_common()])

block("Conteggio per fonte", ["fonte", "n. domande"],
      [[k, v] for k, v in by_fonte.most_common()])

block("Copertura per sotto-argomento", ["dominio", "sotto_argomento", "n. domande"],
      [[d, s, n] for (d, s), n in sorted(by_sub.items(), key=lambda x: (DOMAINS.index(x[0][0]), -x[1]))])

block("Domande rimosse per motivo", ["motivo", "n."],
      [[k, v] for k, v in by_removed.most_common()] or [["(nessuna)", 0]])

for col, w in zip("ABCDE", (52, 46, 16, 18, 14)):
    ws.column_dimensions[col].width = w

# --- Rimosse
ws = wb.create_sheet("Rimosse")
ws.append(["id", "testo_breve", "motivo"])
style_header(ws, 3)
for r in removed:
    ws.append([r.get("id", ""), r["testo_breve"], r["motivo"]])
if not removed:
    ws.append(["", "(nessuna domanda rimossa)", ""])
for col, w in zip("ABC", (12, 100, 26)):
    ws.column_dimensions[col].width = w

# --- Fonti
ws = wb.create_sheet("Fonti")
ws.append(["url", "n_domande_raccolte", "data_raccolta", "note_accessibilita"])
style_header(ws, 4)
src_counts = Counter(q["url_fonte"] for q in final if not q["is_generated"])
fonti_rows = [
    [STUDY_GUIDE_URL, 0, BUILD_DATE,
     "Accessibile. Ancoraggio ufficiale: skills measured al 17 aprile 2026. Usata per domini, pesi e sotto-argomenti."],
    ["https://learn.microsoft.com/en-us/credentials/certifications/exams/az-104/practice/assessment", 0, BUILD_DATE,
     "Pagina applicativa JS: contenuto non estraibile staticamente. Usata solo per calibrare stile e formati, mai per copiare (come da brief)."],
    ["https://learn.microsoft.com/en-us/azure/", int(sum(1 for q in final if q["is_generated"])), BUILD_DATE,
     "Accessibile. Documentazione Microsoft Learn: base di ancoraggio e verifica per tutte le domande generate."],
    ["https://www.examtopics.com/exams/microsoft/az-104/", 0, BUILD_DATE,
     "ESCLUSA DELIBERATAMENTE. Ospita item riprodotti da esami reali (braindump) in violazione dell'NDA Microsoft; "
     "la parafrasi non sana il problema perche il contenuto protetto e lo scenario d'esame, non la sua formulazione. "
     "Rischio di decertificazione per chi studia su questo materiale."],
]
for u, n in src_counts.most_common():
    fonti_rows.append([u, n, BUILD_DATE, "Accessibile. Licenza aperta (MIT): riuso consentito con attribuzione."])
for r in fonti_rows:
    ws.append(r)
for col, w in zip("ABCD", (62, 22, 16, 105)):
    ws.column_dimensions[col].width = w
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# --- README
ws = wb.create_sheet("README")
readme = [
    ("Banca domande AZ-104 — README", "title"),
    ("", ""),
    (f"Versione study guide di riferimento: {STUDY_GUIDE_VERSION}", ""),
    (f"URL study guide: {STUDY_GUIDE_URL}", ""),
    (f"Data di build: {BUILD_DATE}", ""),
    (f"Totale domande: {total}", ""),
    ("", ""),
    ("Formato esame reale (verificato in Fase 0)", "h"),
    ("Domande: circa 40-60 per sessione. Durata: 100 minuti. Punteggio minimo: 700/1000 (scala 1-1000, non e il 70%).", ""),
    ("Tipologie: multiple choice, multiple response, serie yes/no, hotspot, drag-and-drop, case study.", ""),
    ("La simulazione inclusa usa 50 domande / 100 minuti / soglia 700, rispettando i pesi ufficiali dei domini.", ""),
    ("", ""),
    ("Pesi ufficiali dei domini", "h"),
] + [(f"  {d}: {WEIGHTS[d][0]}-{WEIGHTS[d][1]}%  ->  in banca: {by_dom.get(d,0)} ({round(by_dom.get(d,0)/total*100,1) if total else 0}%)", "")
     for d in DOMAINS] + [
    ("", ""),
    ("Provenienza e integrita", "h"),
    ("Le domande sono ORIGINALI, scritte a partire dalla documentazione pubblica Microsoft Learn, oppure raccolte da", ""),
    ("repository con licenza aperta (MIT) e riformulate. Ogni domanda cita una pagina Learn specifica che ne prova la risposta.", ""),
    ("ExamTopics e gli altri siti di dump NON sono stati usati: contengono item riprodotti da esami reali in violazione", ""),
    ("dell'NDA Microsoft. Parafrasarli non risolve il problema, perche cio che e protetto e lo scenario d'esame e non la", ""),
    ("sua formulazione; inoltre studiare su dump espone a decertificazione. Vedi il foglio 'Fonti' per il dettaglio.", ""),
    ("", ""),
    ("ATTENZIONE — LA VERIFICA (FASE 4) E' INCOMPLETA", "h"),
    (f"Solo {by_stato.get('verificata',0)} domande su {total} sono state ricontrollate contro Microsoft Learn da un", ""),
    ("verificatore indipendente. Le altre sono in stato 'da_rivedere': sono state scritte consultando la", ""),
    ("documentazione, ma NESSUNO ha ancora ricontrollato la chiave di risposta in modo indipendente.", ""),
    ("Motivo: la sessione di build ha esaurito il limite di utilizzo e i 35 agenti verificatori sono stati", ""),
    ("interrotti prima di completare il lavoro. Non e' un giudizio sulla qualita' delle domande.", ""),
    ("Conseguenza pratica: usa la banca per studiare, ma se una risposta ti sembra sbagliata controlla il link", ""),
    ("Learn nel campo url_riferimento prima di darla per buona. Il simulatore permette di correggere la", ""),
    ("risposta e salvare l'override. Per completare la verifica, vedi 'Come RIPRENDERE' qui sotto.", ""),
    ("", ""),
    ("Significato del campo 'stato'", "h"),
    ("  verificata   - la documentazione Microsoft Learn conferma la risposta indicata.", ""),
    ("  contestata   - esistono due letture difendibili; la spiegazione riporta entrambe e indica quella sostenuta.", ""),
    ("  da_rivedere  - non verificabile con certezza: da controllare prima dell'uso in esame simulato.", ""),
    ("", ""),
    ("Campi sintesi_commenti / consenso_community", "h"),
    ("Volutamente vuoti quando non esiste una discussione pubblica legittima sulla domanda. Non sono mai stati inventati:", ""),
    ("un consenso fittizio sarebbe peggio di un campo vuoto perche darebbe falsa fiducia su una risposta.", ""),
    ("", ""),
    ("Come RIPRENDERE il lavoro in una sessione successiva", "h"),
    ("PRIORITA' ASSOLUTA: completare la Fase 4 sulle domande in stato 'da_rivedere' (vedi avviso sopra).", ""),
    ("   Non servono nuove domande: ne servono di verificate. Il target di 500 e' gia' superato.", ""),
    ("1. Allega az104_question_bank.json (o .xlsx) al nuovo prompt e indica il parametro RIPRESA.", ""),
    ("2. Gli id sono progressivi (AZ104-0001...). Riprendi la numerazione dall'ultimo id presente: non rinumerare.", ""),
    ("3. Aggiungi solo domande NUOVE (append). Esegui la deduplica semantica contro le esistenti prima di inserirle", ""),
    ("   (soglia Jaccard 0.72 su trigrammi del campo 'domanda' - stessa logica di questo build).", ""),
    ("4. Ricontrolla l'ancoraggio: se la study guide e stata aggiornata dopo il 17 aprile 2026, rivedi pesi e obsolescenza", ""),
    ("   prima di aggiungere altro.", ""),
    ("5. Priorita di lavoro: prima le domande in stato 'da_rivedere', poi i sotto-argomenti con copertura piu bassa", ""),
    ("   (vedi foglio 'Statistiche' -> 'Copertura per sotto-argomento').", ""),
    ("", ""),
    ("File collegati", "h"),
    ("  az104_question_bank.json  - stesso contenuto del foglio 'Domande'; e il file che alimenta l'app simulatore.", ""),
    ("  Simulatore                - app React: carica il JSON, modalita Studio ed Esame, persistenza via window.storage.", ""),
]
for text, kind in readme:
    ws.append([text])
    if kind == "title":
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=14, color="1F3864")
    elif kind == "h":
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11, color="1F3864")
ws.column_dimensions["A"].width = 125

xlsx_path = os.path.join(OUT_DIR, "az104_question_bank.xlsx")
wb.save(xlsx_path)
print(f"[out] {xlsx_path}")

# ---------------------------------------------------------------- console report
print("\n=== COPERTURA PER DOMINIO ===")
print(f"{'dominio':<46}{'n':>5}{'%':>8}   peso      esito")
for d in DOMAINS:
    n = by_dom.get(d, 0)
    pct = n / total * 100 if total else 0
    lo, hi = WEIGHTS[d]
    esito = "OK" if lo <= pct <= hi else ("SOTTO" if pct < lo else "SOPRA")
    print(f"{d:<46}{n:>5}{pct:>7.1f}%   {lo}-{hi}%     {esito}")
print(f"{'TOTALE':<46}{total:>5}")
print("\nstato:", dict(by_stato))
print("tipo:", dict(by_tipo))
print("difficolta:", dict(sorted(by_diff.items())))
print("origine:", dict(by_gen))
print("rimosse:", dict(by_removed), "tot", len(removed))

with open(os.path.join(BASE, "build_stats.json"), "w", encoding="utf-8") as fh:
    json.dump({
        "total": total,
        "by_domain": dict(by_dom),
        "by_stato": dict(by_stato),
        "by_tipo": dict(by_tipo),
        "by_difficolta": {str(k): v for k, v in by_diff.items()},
        "by_fonte": dict(by_fonte),
        "by_origine": dict(by_gen),
        "removed": dict(by_removed),
        "removed_total": len(removed),
        "by_subtopic": {f"{d} :: {s}": n for (d, s), n in by_sub.items()},
        "load_report": [{"batch": b, "n": n, "status": s} for b, n, s in load_report],
    }, fh, ensure_ascii=False, indent=1)
print("\n[out] build_stats.json")
