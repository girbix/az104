#!/usr/bin/env python3
"""
Builds the Italian bank from the translated chunks, but only after proving the
translation did not touch anything the simulator treats as a key.

Fails loudly rather than shipping a bank that silently mis-grades.
"""
import json, os, re, sys, glob
from collections import Counter

BASE = os.path.dirname(os.path.abspath(__file__))
TX = os.path.join(BASE, "tx")
OUT_DIR = sys.argv[1] if len(sys.argv) > 1 else BASE
EN_PATH = os.path.join(OUT_DIR, "az104_question_bank.json")

BUILD_DATE = "2026-07-16"
STUDY_GUIDE_VERSION = "Skills measured as of April 17, 2026"
STUDY_GUIDE_URL = "https://learn.microsoft.com/en-us/credentials/certifications/resources/study-guides/az-104"

DOMAINS = [
    "Manage Azure identities and governance",
    "Implement and manage storage",
    "Deploy and manage Azure compute resources",
    "Implement and manage virtual networking",
    "Monitor and maintain Azure resources",
]
WEIGHTS = {
    "Manage Azure identities and governance": (20, 25),
    "Implement and manage storage": (15, 20),
    "Deploy and manage Azure compute resources": (20, 25),
    "Implement and manage virtual networking": (15, 20),
    "Monitor and maintain Azure resources": (10, 15),
}
FIELDS = ["id", "dominio", "sotto_argomento", "tipo", "domanda",
          "opzione_a", "opzione_b", "opzione_c", "opzione_d", "opzione_e",
          "risposta_corretta", "spiegazione", "url_riferimento", "fonte", "url_fonte",
          "sintesi_commenti", "consenso_community", "stato", "difficolta", "tags",
          "is_generated", "data_verifica"]

# Fields the translator was told never to touch. These are what the app matches on.
FROZEN = ["id", "dominio", "sotto_argomento", "tipo", "risposta_corretta",
          "url_riferimento", "fonte", "url_fonte", "stato", "difficolta",
          "is_generated", "data_verifica"]
TRANSLATED = ["domanda", "opzione_a", "opzione_b", "opzione_c", "opzione_d",
              "opzione_e", "spiegazione"]

en = {q["id"]: q for q in json.load(open(EN_PATH, encoding="utf-8"))}
print(f"[en] {len(en)} domande di riferimento")

it = {}
missing_files = []
for i in range(1, 15):
    p = os.path.join(TX, f"chunk_{i:02d}.it.json")
    if not os.path.exists(p):
        missing_files.append(f"chunk_{i:02d}.it.json")
        continue
    try:
        txt = open(p, encoding="utf-8").read().strip()
        txt = re.sub(r"^```(?:json)?\s*|\s*```$", "", txt)
        for q in json.loads(txt):
            if isinstance(q, dict) and q.get("id"):
                it[q["id"]] = q
    except Exception as e:
        print(f"  !! {os.path.basename(p)} illeggibile: {type(e).__name__}: {e}")

print(f"[it] {len(it)} domande tradotte da {14 - len(missing_files)}/14 lotti")
if missing_files:
    print(f"  !! lotti mancanti: {', '.join(missing_files)}")

# ------------------------------------------------------------------ validate
errors, warns = [], []

for qid, e in en.items():
    t = it.get(qid)
    if not t:
        errors.append(f"{qid}: manca la traduzione")
        continue
    # 1. frozen fields must be byte-identical
    for f in FROZEN:
        a, b = e.get(f), t.get(f)
        if isinstance(a, bool) or isinstance(b, bool):
            same = bool(a) == bool(b)
        elif isinstance(a, int) or isinstance(b, int):
            same = str(a) == str(b)
        else:
            same = str(a or "") == str(b or "")
        if not same:
            errors.append(f"{qid}: campo bloccato '{f}' modificato: {a!r} -> {b!r}")
    # 2. hotspot dropdown choices must be identical, or grading breaks
    if e.get("tipo") == "hotspot":
        for c in "abcde":
            oe, ot = str(e.get("opzione_" + c) or ""), str(t.get("opzione_" + c) or "")
            if not oe:
                continue
            me = re.search(r"\[(.*)\]\s*$", oe, re.S)
            mt = re.search(r"\[(.*)\]\s*$", ot, re.S)
            if not mt:
                errors.append(f"{qid}: hotspot opzione_{c} ha perso il blocco [..]")
            elif me and me.group(1) != mt.group(1):
                errors.append(f"{qid}: hotspot opzione_{c} valori alterati: "
                              f"{me.group(1)[:45]!r} -> {mt.group(1)[:45]!r}")
    # 3. an empty option must stay empty and vice versa (shape must match)
    for c in "abcde":
        if bool(str(e.get("opzione_" + c) or "").strip()) != bool(str(t.get("opzione_" + c) or "").strip()):
            errors.append(f"{qid}: opzione_{c} presenza/assenza non coerente")
    # 4. did it actually translate? (heuristic, warning only)
    if str(e.get("domanda") or "").strip() == str(t.get("domanda") or "").strip():
        warns.append(f"{qid}: domanda identica all'inglese (non tradotta?)")

extra = set(it) - set(en)
if extra:
    warns.append(f"{len(extra)} id tradotti non presenti nella banca EN (ignorati)")

print(f"\n[validazione] errori: {len(errors)}  avvisi: {len(warns)}")
for x in errors[:15]:
    print("   ERRORE", x)
if len(errors) > 15:
    print(f"   ... e altri {len(errors)-15}")
for x in warns[:8]:
    print("   avviso", x)
if len(warns) > 8:
    print(f"   ... e altri {len(warns)-8}")

if errors:
    print("\nBUILD INTERROTTA: la banca italiana non viene scritta finche' gli errori non sono risolti.")
    sys.exit(1)
if missing_files:
    print("\nBUILD INTERROTTA: mancano lotti di traduzione.")
    sys.exit(1)

# ------------------------------------------------------------------ assemble
final = []
for qid in [q["id"] for q in json.load(open(EN_PATH, encoding="utf-8"))]:
    e, t = en[qid], it[qid]
    o = {}
    for f in FIELDS:
        o[f] = t.get(f) if f in TRANSLATED else e.get(f)
        if f not in ("difficolta", "is_generated"):
            o[f] = "" if o[f] is None else str(o[f])
    o["difficolta"] = int(e["difficolta"])
    o["is_generated"] = bool(e["is_generated"])
    final.append(o)

json_path = os.path.join(OUT_DIR, "az104_question_bank_it.json")
json.dump(final, open(json_path, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print(f"\n[out] {json_path}  ({len(final)} domande)")

# ------------------------------------------------------------------ xlsx
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

total = len(final)
by_dom = Counter(q["dominio"] for q in final)
by_stato = Counter(q["stato"] for q in final)
by_tipo = Counter(q["tipo"] for q in final)
by_diff = Counter(q["difficolta"] for q in final)
by_sub = Counter((q["dominio"], q["sotto_argomento"]) for q in final)

wb = Workbook()
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=12, color="1F3864")


def style_header(ws, n, row=1):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


ws = wb.active
ws.title = "Domande"
ws.append(FIELDS)
for q in final:
    ws.append([q[f] for f in FIELDS])
style_header(ws, len(FIELDS))
ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}{ws.max_row}"
widths = {"id": 12, "dominio": 34, "sotto_argomento": 34, "tipo": 17, "domanda": 70,
          "opzione_a": 30, "opzione_b": 30, "opzione_c": 30, "opzione_d": 30, "opzione_e": 30,
          "risposta_corretta": 16, "spiegazione": 80, "url_riferimento": 45, "fonte": 30,
          "url_fonte": 40, "sintesi_commenti": 40, "consenso_community": 18, "stato": 13,
          "difficolta": 10, "tags": 28, "is_generated": 12, "data_verifica": 13}
for i, f in enumerate(FIELDS, 1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(f, 18)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

ws = wb.create_sheet("Statistiche")
ws.append(["Banca domande AZ-104 (IT) — Statistiche"])
ws["A1"].font = Font(bold=True, size=14, color="1F3864")
ws.append([f"Riferimento: {STUDY_GUIDE_VERSION}"])
ws.append([f"Data build: {BUILD_DATE}"])
ws.append([f"Totale domande: {total}"])


def block(title, headers, rows):
    ws.append([])
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=title).font = TITLE_FONT
    ws.append(headers)
    style_header(ws, len(headers), row=ws.max_row)
    for row in rows:
        ws.append(row)
    ws.append([])


rows = []
for d in DOMAINS:
    n = by_dom.get(d, 0)
    p = n / total * 100 if total else 0
    lo, hi = WEIGHTS[d]
    rows.append([d, n, round(p, 1), f"{lo}-{hi}%", "OK" if lo <= p <= hi else "FUORI BANDA"])
block("Copertura per dominio vs pesi ufficiali",
      ["dominio", "n. domande", "% attuale", "peso ufficiale", "esito"], rows)
block("Conteggio per stato", ["stato", "n.", "%"],
      [[k, v, round(v / total * 100, 1)] for k, v in by_stato.most_common()])
block("Conteggio per tipo", ["tipo", "n.", "%"],
      [[k, v, round(v / total * 100, 1)] for k, v in by_tipo.most_common()])
block("Conteggio per difficolta", ["difficolta", "n.", "%"],
      [[f"{k} ({['','base','applicativa','scenario complesso'][k]})", v, round(v / total * 100, 1)]
       for k, v in sorted(by_diff.items())])
block("Copertura per sotto-argomento", ["dominio", "sotto_argomento", "n."],
      [[d, s, n] for (d, s), n in sorted(by_sub.items(), key=lambda x: (DOMAINS.index(x[0][0]), -x[1]))])
for col, w in zip("ABCDE", (52, 46, 16, 18, 14)):
    ws.column_dimensions[col].width = w

ws = wb.create_sheet("README")
readme = [
    ("Banca domande AZ-104 — VERSIONE ITALIANA", "title"),
    ("", ""),
    ("Questa e' la traduzione italiana della banca. La versione inglese e' in az104_question_bank.xlsx", ""),
    ("/ az104_question_bank.json ed e' rimasta invariata: le due banche sono allineate id per id", ""),
    ("(AZ104-0001 qui = AZ104-0001 nella versione inglese).", ""),
    ("", ""),
    ("Quale versione usare", "h"),
    ("L'esame reale si sostiene in inglese o in una lingua localizzata. Se pensi di sostenerlo in inglese,", ""),
    ("allenati sulla versione inglese: il lessico dei distrattori fa parte della difficolta' della prova.", ""),
    ("Usa questa versione italiana per capire i concetti piu' in fretta, non come unica preparazione.", ""),
    ("", ""),
    ("Cosa e' stato tradotto e cosa no", "h"),
    ("Tradotti:     domanda, opzione_a..opzione_e, spiegazione", ""),
    ("NON tradotti: id, dominio, sotto_argomento, tipo, risposta_corretta, url_riferimento, fonte,", ""),
    ("              url_fonte, stato, difficolta, tags, is_generated, data_verifica", ""),
    ("Motivo: sono chiavi tecniche che il simulatore confronta con stringhe esatte (l'estrazione pesata per", ""),
    ("dominio e la valutazione delle risposte). Tradurle romperebbe l'app. 'dominio' resta quindi il nome", ""),
    ("ufficiale inglese dell'obiettivo, ma l'interfaccia del simulatore lo mostra tradotto.", ""),
    ("", ""),
    ("Restano in inglese anche dentro il testo tradotto i nomi di servizio Azure (Microsoft Entra ID,", ""),
    ("Azure Blob Storage...), i ruoli (Owner, Contributor...), gli SKU (Standard_LRS, Premium SSD v2...),", ""),
    ("i cmdlet e il codice: sono gli stessi termini che troverai nel portale e nell'esame.", ""),
    ("Nelle domande hotspot i valori dei menu a discesa sono volutamente NON tradotti: devono combaciare", ""),
    ("carattere per carattere con risposta_corretta, altrimenti la domanda diventa impossibile da valutare.", ""),
    ("", ""),
    (f"ATTENZIONE — LA VERIFICA (FASE 4) E' INCOMPLETA", "h"),
    (f"Solo {by_stato.get('verificata',0)} domande su {total} sono state ricontrollate contro Microsoft Learn da un", ""),
    ("verificatore indipendente. Le altre sono in stato 'da_rivedere': scritte consultando la documentazione,", ""),
    ("ma con la chiave di risposta non ancora ricontrollata in modo indipendente. La sessione di build ha", ""),
    ("esaurito il limite di utilizzo e i verificatori sono stati interrotti. Se una risposta ti sembra", ""),
    ("sbagliata, controlla il link Learn in url_riferimento prima di darla per buona.", ""),
    ("", ""),
    ("La traduzione NON e' una verifica: un errore tecnico presente nell'originale inglese e' stato tradotto", ""),
    ("fedelmente e si trova identico anche qui.", ""),
    ("", ""),
    ("Significato del campo 'stato'", "h"),
    ("  verificata   - la documentazione Microsoft Learn conferma la risposta indicata.", ""),
    ("  contestata   - esistono due letture difendibili; la spiegazione riporta entrambe.", ""),
    ("  da_rivedere  - non verificabile con certezza: da controllare prima di darla per buona.", ""),
    ("", ""),
    ("Come RIPRENDERE il lavoro", "h"),
    ("PRIORITA': completare la Fase 4 sulle domande 'da_rivedere'. Il target di 500 e' gia' superato:", ""),
    ("non servono nuove domande, servono domande verificate.", ""),
    ("Se correggi una risposta, correggila in ENTRAMBE le banche (stesso id) per non farle divergere.", ""),
    ("", ""),
    ("File collegati", "h"),
    ("  az104_question_bank_it.json  - questa banca, in JSON: caricala nel simulatore per studiare in italiano.", ""),
    ("  az104_question_bank.json     - la stessa banca in inglese, allineata id per id.", ""),
]
for text, kind in readme:
    ws.append([text])
    if kind == "title":
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=14, color="1F3864")
    elif kind == "h":
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11, color="1F3864")
ws.column_dimensions["A"].width = 125

xlsx_path = os.path.join(OUT_DIR, "az104_question_bank_it.xlsx")
wb.save(xlsx_path)
print(f"[out] {xlsx_path}")

print("\n=== COPERTURA (IT) ===")
for d in DOMAINS:
    n = by_dom.get(d, 0)
    print(f"{d:<46}{n:>5}{n/total*100:>7.1f}%")
print(f"{'TOTALE':<46}{total:>5}")
print("stato:", dict(by_stato))
